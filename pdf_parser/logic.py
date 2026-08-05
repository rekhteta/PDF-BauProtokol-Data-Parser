import os
import logging
import re
import concurrent.futures
import numpy as np
import pandas as pd
import fitz  # PyMuPDF
from typing import Callable, List, Dict, Any, Optional

from .config import load_profiles

class PDFParserLogic:
    """
    Business Logic for extracting data from PDF files.
    Routes to specific protocol extractors based on PDF text.
    Handles parallel execution and numpy-optimized visual element validation.
    """
    def __init__(self, target_folder: str, output_excel_path: str, progress_callback: Callable, finish_callback: Callable, selected_columns: List[str]):
        self.target_folder = target_folder
        self.output_excel_path = output_excel_path
        self.progress_callback = progress_callback
        self.finish_callback = finish_callback
        self.selected_columns = selected_columns

    def run(self):
        try:
            logging.info("Starting PDF analysis run...")
            self._analyze()
        except Exception as e:
            logging.error("Fatal error during PDF analysis", exc_info=True)
            self.finish_callback(False, str(e), None)
            
    def _analyze(self):
        self.progress_callback("Identifying PDF files...", 0, 0)
        
        pdf_files = []
        for root, dirs, files in os.walk(self.target_folder):
            for f in files:
                if f.lower().endswith('.pdf'):
                    pdf_files.append(os.path.join(root, f))
            
        total_items = len(pdf_files)
        logging.info(f"Found {total_items} PDF files in target directory '{self.target_folder}'")
        
        if total_items == 0:
            logging.warning("No PDF files found.")
            self.finish_callback(False, "No PDF files found in the selected folder.", None)
            return

        processed_count = 0
        data = []
        
        # Concurrency level: auto-detect based on CPU cores
        max_workers = min(32, (os.cpu_count() or 1) + 4)
        logging.info(f"Processing PDFs with ThreadPoolExecutor (max_workers={max_workers})")

        profiles = load_profiles()

        def process_file(file_path: str) -> Optional[Dict[str, Any]]:
            file_name = os.path.basename(file_path)
            logging.info(f"Start parsing file: {file_name}")
            try:
                with fitz.open(file_path) as doc:
                    if len(doc) == 0:
                        logging.warning(f"File {file_name} has 0 pages. Skipping.")
                        return None
                        
                    page0 = doc[0]
                    page0_text = page0.get_text("text")
                    
                    # Manual edits detection
                    has_annots = any(True for _ in page0.annots())
                    has_widgets = any(True for _ in page0.widgets())
                    
                    # Routing Logic using profiles loaded dynamically
                    matched_profile = None
                    for p in profiles:
                        # If trigger_keywords matches any keyword in page0_text
                        if any(kw.lower() in page0_text.lower() for kw in p.get("trigger_keywords", [])):
                            matched_profile = p
                            break
                            
                    extraction_result = {}
                    extraction_result["Has Annotations"] = has_annots
                    extraction_result["Has Form Fields"] = has_widgets
                    
                    if matched_profile:
                        logging.info(f"File {file_name} matched profile '{matched_profile['name']}'")
                        extraction_result["Protocol Type"] = matched_profile["name"]
                        
                        # Cache for hausanschluss extractor to avoid parsing multiple times
                        hausanschluss_cache = None
                        
                        for col_name, rules in matched_profile["columns"].items():
                            val = ""
                            for rule in rules:
                                rtype = rule.get("type")
                                if rtype == "inline":
                                    from .extractors import InlineExtractor
                                    val = InlineExtractor(*rule.get("keys", [])).extract(page0_text)
                                elif rtype == "next_line":
                                    from .extractors import NextLineExtractor
                                    val = NextLineExtractor(*rule.get("keys", [])).extract(page0_text)
                                elif rtype == "fixed_index":
                                    from .extractors import PositionalExtractor
                                    val = PositionalExtractor(rule.get("index", 0)).extract(page0_text)
                                elif rtype == "speednet_positional":
                                    from .extractors import SpeedNetPositionalExtractor
                                    val = SpeedNetPositionalExtractor(rule.get("index", 0), rule.get("pattern", "")).extract(page0_text)
                                elif rtype == "regex":
                                    pattern = rule.get("pattern", "")
                                    group = rule.get("group", 1)
                                    pick_last = rule.get("pick", "") == "last"
                                    if pattern:
                                        matches = list(re.finditer(pattern, page0_text, re.IGNORECASE | re.MULTILINE))
                                        if matches:
                                            m = matches[-1] if pick_last else matches[0]
                                            try:
                                                val = m.group(group).strip()
                                            except IndexError:
                                                val = m.group(0).strip()
                                elif rtype == "regex_combine":
                                    patterns = rule.get("patterns", [])
                                    fmt = rule.get("format", "{0}")
                                    captured = []
                                    for pat in patterns:
                                        m = re.search(pat, page0_text, re.IGNORECASE | re.MULTILINE)
                                        captured.append(m.group(1).strip() if m else "")
                                    if any(captured):
                                        try:
                                            val = fmt.format(*captured)
                                        except (IndexError, KeyError):
                                            val = " / ".join(c for c in captured if c)
                                elif rtype == "special":
                                    name = rule.get("name")
                                    if name == "strecke_extractor":
                                        from .extractors import StreckeExtractor
                                        val = StreckeExtractor().extract(page0_text)
                                    elif name == "meterzahlen_extractor":
                                        from .extractors import MeterzahlenExtractor
                                        val = MeterzahlenExtractor().extract(page0_text)
                                    elif name.startswith("hausanschluss_"):
                                        if hausanschluss_cache is None:
                                            hausanschluss_cache = self._extract_hausanschluss(doc, page0_text)
                                        
                                        key_map = {
                                            "hausanschluss_bbnd_id": "BBND ID",
                                            "hausanschluss_anzahl_we": "Anzahl WE",
                                            "hausanschluss_nvt": "Bezeichnung NVt",
                                            "hausanschluss_date": "Datum Herstellung Hausanschluss",
                                            "hausanschluss_rohrverband": "Bezeichnung Rohrverband",
                                            "hausanschluss_farbe": "Farbe",
                                            "hausanschluss_verbundrohr": "Verbundrohr",
                                            "hausanschluss_bild1": "Bild 1",
                                            "hausanschluss_bild2": "Bild 2",
                                            "hausanschluss_bild3": "Bild 3",
                                            "hausanschluss_bild4": "Bild 4",
                                            "hausanschluss_signature": "Unterschrift"
                                        }
                                        val = hausanschluss_cache.get(key_map.get(name, ""), "")
                                if val:
                                    break
                            extraction_result[col_name] = val
                            
                        # Post-processing normalization
                        
                        # 1. Date normalization for Einblasdatum
                        if "Einblasdatum" in extraction_result:
                            raw_date = extraction_result["Einblasdatum"]
                            if raw_date and not re.search(r"\d{1,2}[.\-/: ]\d{1,2}[.\-/: ]\d{2,4}", raw_date):
                                date_match = re.search(
                                    r"(\d{1,2}[.\-/: ]\d{1,2}[.\-/: ]\d{2,4})\s+(\d{1,2}:\d{2}(?::\d{2})?)",
                                    page0_text
                                )
                                if date_match:
                                    raw_date = date_match.group(0)
                                else:
                                    raw_date = ""
                            extraction_result["Einblasdatum"] = self._normalize_date(raw_date)
                            
                        # 2. Streckenabschnitt Start/Ziel split
                        if "Streckenabschnitt" in extraction_result:
                            sa_text = extraction_result["Streckenabschnitt"]
                            start_val, ziel_val = "", ""
                            if sa_text:
                                clean_text = sa_text.replace('_', ' ').strip()
                                parts = clean_text.split(' ', 1)
                                if len(parts) >= 1:
                                    start_val = parts[0].upper()
                                if len(parts) > 1:
                                    remainder = parts[1].strip()
                                    prep_match = re.match(
                                        r'^(?:zu|in|nach|bis|an|richtung)\s+(.+)$',
                                        remainder, re.IGNORECASE
                                    )
                                    ziel_val = prep_match.group(1).upper() if prep_match else remainder.upper()
                            extraction_result["Start"] = start_val
                            extraction_result["Ziel"] = ziel_val
                            
                        # 3. Clean up Baumaße if it contains the label
                        if "Baumaße:" in extraction_result:
                            if "Baumaße:" in extraction_result["Baumaße:"]:
                                extraction_result["Baumaße:"] = extraction_result["Baumaße:"].replace("Baumaße:", "").strip()
                                
                    else:
                        logging.info(f"File {file_name} has unknown format layout")
                        extraction_result["Protocol Type"] = "Unknown Format"
                        
                    # Add metadata
                    folder_path_full = os.path.dirname(file_path)
                    extraction_result["File Name"] = file_name
                    extraction_result["Full Path"] = file_path
                    extraction_result["Folder Path"] = folder_path_full
                    extraction_result["Folder Name"] = os.path.basename(folder_path_full)
                    
                    # File Size
                    try:
                        size_bytes = os.path.getsize(file_path)
                        extraction_result["File Size"] = f"{round(size_bytes / 1024, 1)} KB"
                    except Exception:
                        extraction_result["File Size"] = ""
                        
                    # Created and Modified Dates
                    from datetime import datetime
                    try:
                        c_time = os.path.getctime(file_path)
                        extraction_result["Created Date"] = datetime.fromtimestamp(c_time).strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        extraction_result["Created Date"] = ""
                        
                    try:
                        m_time = os.path.getmtime(file_path)
                        extraction_result["Last Modify Date"] = datetime.fromtimestamp(m_time).strftime("%Y-%m-%d %H:%M:%S")
                    except Exception:
                        extraction_result["Last Modify Date"] = ""
                        
                    # Created By (File Owner)
                    owner_name = "Unknown"
                    try:
                        import win32security
                        sd = win32security.GetFileSecurity(file_path, win32security.OWNER_SECURITY_INFORMATION)
                        owner_sid = sd.GetSecurityDescriptorOwner()
                        name, domain, type = win32security.LookupAccountSid(None, owner_sid)
                        owner_name = f"{domain}\\{name}"
                    except Exception:
                        try:
                            import subprocess
                            cmd = f'powershell -Command "(Get-Acl \'{file_path}\').Owner"'
                            res = subprocess.check_output(cmd, shell=True, text=True, timeout=2).strip()
                            if res:
                                owner_name = res
                        except Exception:
                            pass
                    extraction_result["Created By"] = owner_name

                    return extraction_result
            except Exception as e:
                logging.error(f"Error parsing PDF '{file_name}': {e}", exc_info=True)
                return {
                    "File Name": file_name,
                    "ERROR": f"Failed to read content: {e}",
                    "Full Path": file_path,
                    "Has Annotations": False,
                    "Has Form Fields": False
                }

        # Run concurrent extraction
        with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
            futures = {executor.submit(process_file, path): path for path in pdf_files}
            for future in concurrent.futures.as_completed(futures):
                result = future.result()
                if result:
                    data.append(result)
                processed_count += 1
                if processed_count % 5 == 0 or processed_count == total_items:
                    self.progress_callback(f"Processed {processed_count} of {total_items} PDFs", processed_count, total_items)

        # Save to Excel
        try:
            if not data:
                logging.warning("No data extracted from PDFs.")
                self.finish_callback(False, "No data extracted.", None)
                return
                
            logging.info(f"Formatting extracted data into Excel. Total rows: {len(data)}")
            df = pd.DataFrame(data)
            
            # Filter and reorder based on selected columns
            all_possible_cols = df.columns.tolist()
            final_cols = []
            for c in self.selected_columns:
                if c in all_possible_cols and c not in final_cols:
                    final_cols.append(c)
            
            if not final_cols:
                logging.warning("None of the selected columns were found in data; falling back to all columns.")
                final_cols = all_possible_cols
            
            df = df[final_cols]
            
            # Coerce Strecke to numeric (integer where possible, else float)
            if "Strecke" in df.columns:
                df["Strecke"] = pd.to_numeric(df["Strecke"], errors="coerce")
                # Use Int64 (nullable integer) so cells with valid integers show as int, not float
                try:
                    df["Strecke"] = df["Strecke"].round(3)
                    if df["Strecke"].dropna().apply(lambda x: x == int(x)).all():
                        df["Strecke"] = df["Strecke"].astype("Int64")
                except Exception:
                    pass

            self._write_formatted_excel(df, self.output_excel_path)
            logging.info(f"Excel file successfully saved to '{self.output_excel_path}'")
            self.finish_callback(True, "Success", self.output_excel_path)
        except Exception as e:
            logging.error("Failed writing Excel file", exc_info=True)
            self.finish_callback(False, f"Target file might be open or read-only:\n{str(e)}", None)

    def _write_formatted_excel(self, df: pd.DataFrame, path: str):
        """Writes DataFrame to Excel with formatted headers: centered, wrap text, no borders.
        Also converts Einblasdatum to proper datetime and applies German date format."""
        from openpyxl import load_workbook
        from openpyxl.styles import Alignment, Border, Side
        from datetime import datetime

        # Convert Einblasdatum string -> datetime so Excel stores a real date value
        if "Einblasdatum" in df.columns:
            def _parse_dt(val):
                if not val or not isinstance(val, str):
                    return val
                for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d %H:%M", "%Y-%m-%d"):
                    try:
                        return datetime.strptime(val.strip(), fmt)
                    except ValueError:
                        continue
                return val
            df = df.copy()
            df["Einblasdatum"] = df["Einblasdatum"].apply(_parse_dt)

        df.to_excel(path, index=False)
        wb = load_workbook(path)
        ws = wb.active

        no_border = Border(
            left=Side(border_style=None),
            right=Side(border_style=None),
            top=Side(border_style=None),
            bottom=Side(border_style=None)
        )
        header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Find the column index for Einblasdatum (1-based)
        date_col_idx = None
        if "Einblasdatum" in df.columns:
            date_col_idx = df.columns.tolist().index("Einblasdatum") + 1  # +1 for 1-based

        for cell in ws[1]:  # Row 1 = headers
            cell.alignment = header_align
            cell.border = no_border

        # Apply German date format JJ.MM.TT hh:mm to all data cells in the Einblasdatum column
        if date_col_idx is not None:
            for row in ws.iter_rows(min_row=2, min_col=date_col_idx, max_col=date_col_idx):
                for cell in row:
                    if cell.value:
                        cell.number_format = "YY.MM.DD HH:MM"

        wb.save(path)

    def _normalize_date(self, raw_date: str) -> str:
        """Standardizes inconsistent date formats into YYYY-MM-DD HH:MM:SS."""
        if not raw_date:
            return ""
        
        # 1. Split date and time (handling spaces, commas, or specific characters)
        raw_date = raw_date.replace("\n", " ").strip()
        dt_parts = re.split(r"[\s,]+", raw_date)
        if not dt_parts:
            return raw_date
            
        date_part = dt_parts[0]
        time_part = dt_parts[1] if len(dt_parts) > 1 else ""
        
        # 2. Normalize DATE part (DD.MM.YYYY or YYYY-MM-DD)
        date_part = re.sub(r"[:\-]", ".", date_part)
        
        # Check if it's European DD.MM.YYYY
        euro_match = re.match(r"(\d{1,2})\.(\d{1,2})\.(\d{2,4})", date_part)
        if euro_match:
            d, m, y = euro_match.groups()
            if len(y) == 2: y = "20" + y # 26 -> 2026
            date_part = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        else:
            # Check if it's already YYYY.MM.DD
            iso_match = re.match(r"(\d{4})\.(\d{1,2})\.(\d{1,2})", date_part)
            if iso_match:
                y, m, d = iso_match.groups()
                date_part = f"{y}-{m.zfill(2)}-{d.zfill(2)}"
        
        # 3. Normalize TIME part
        if time_part:
            time_part = re.sub(r"[.\-]", ":", time_part)
            t_parts = time_part.split(":")
            if len(t_parts) >= 2:
                h = t_parts[0].zfill(2)
                mn = t_parts[1].zfill(2)
                s = t_parts[2].zfill(2) if len(t_parts) > 2 else "00"
                time_part = f"{h}:{mn}:{s}"

        normalized = date_part + (f" {time_part}" if time_part else " 00:00:00")
        
        try:
            from dateutil import parser
            dt = parser.parse(normalized)
            return dt.strftime("%Y-%m-%d %H:%M:%S")
        except Exception:
            return normalized

    def _extract_hausanschluss(self, doc, text_p0):
        results = {}
        try:
            page0 = doc[0]
            widgets = {w.field_name: w.field_value for w in page0.widgets()}
            
            results["BBND ID"] = widgets.get("ID", "")
            results["Anzahl WE"] = widgets.get("Dropdown1", "")
            results["Bezeichnung NVt"] = widgets.get("Text9", "")
            results["Datum Herstellung Hausanschluss"] = widgets.get("Date1", "")
            results["Bezeichnung Rohrverband"] = widgets.get("Text10", "")
            
            # Farbe Logic
            found_farbe = ""
            active_color_row_y = -1
            for w in page0.widgets():
                if w.field_type == fitz.PDF_WIDGET_TYPE_CHECKBOX:
                     if w.field_value in ["Yes", "On", "True", "1"]:
                          if w.rect.x0 > 450:
                               active_color_row_y = (w.rect.y0 + w.rect.y1) / 2
                               break
                               
            if active_color_row_y != -1:
                search_r = fitz.Rect(400, active_color_row_y - 10, 600, active_color_row_y + 10)
                txt_in_row = page0.get_text("text", clip=search_r).lower()
                colors = ["grau", "orange", "braun", "schwarz", "blau", "rot", "gelb", "weiß", "weiss", "violett", "türkis", "rosa"]
                for c in colors:
                    if c in txt_in_row:
                        found_farbe = c
                        break
            results["Farbe"] = found_farbe if found_farbe else "grau"
            
            # Verbundrohr Logic
            verbund_match = re.search(r"(\d+x\d+(?:mm)?)", text_p0)
            results["Verbundrohr"] = verbund_match.group(1) if verbund_match else "12x10mm"

            # Image & Signature Checks - Optimized using np.frombuffer to avoid heavy list conversions
            zones = {
                 "Bild 1": fitz.Rect(60, 335, 310, 520),
                 "Bild 2": fitz.Rect(315, 335, 555, 520),
                 "Bild 3": fitz.Rect(60, 545, 310, 730),
                 "Bild 4": fitz.Rect(315, 545, 555, 730),
                 "Unterschrift": fitz.Rect(170, 750, 310, 775)
            }
            img_results = {}
            for key, rect in zones.items():
                if rect.x1 > page0.rect.width: rect.x1 = page0.rect.width
                if rect.y1 > page0.rect.height: rect.y1 = page0.rect.height
                
                sub_pix = page0.get_pixmap(clip=rect, dpi=72)
                
                # NumPy performance optimization: use frombuffer instead of list conversion
                if not sub_pix.samples:
                    has_content = False
                else:
                    raw_bytes = np.frombuffer(sub_pix.samples, dtype=np.uint8)
                    std_dev = np.std(raw_bytes)
                    if key == "Unterschrift":
                        has_content = std_dev > 25.0 
                    else:
                        has_content = std_dev > 45.0
                img_results[key] = has_content

            results["Bild 1"] = img_results["Bild 1"]
            results["Bild 2"] = img_results["Bild 2"]
            results["Bild 3"] = img_results["Bild 3"]
            results["Bild 4"] = img_results["Bild 4"]
            results["Unterschrift"] = img_results["Unterschrift"]
            
        except Exception as e:
            results["Extractor Exception"] = str(e)
            
        return results
